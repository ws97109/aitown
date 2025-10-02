"""
問卷結果匯出器
支援將問卷回應匯出為多種格式：CSV、JSON、Excel等
"""

import csv
import json
import os
from datetime import datetime
from typing import List, Dict, Any, Optional
from .models import Survey, SurveyResponse, SurveyManager


class BaseExporter:
    """基礎匯出器類別"""
    
    def __init__(self, survey_manager: SurveyManager):
        self.survey_manager = survey_manager
    
    def export(self, survey_id: str, output_path: str) -> bool:
        """匯出問卷結果的抽象方法"""
        raise NotImplementedError


class CSVExporter(BaseExporter):
    """CSV格式匯出器"""
    
    def export(self, survey_id: str, output_path: str = None) -> str:
        """匯出問卷回應為CSV格式"""
        survey = self.survey_manager.load_survey(survey_id)
        if not survey:
            raise ValueError(f"問卷 {survey_id} 不存在")
        
        responses = self.survey_manager.get_responses_by_survey(survey_id)
        if not responses:
            raise ValueError(f"問卷 {survey_id} 沒有回應數據")
        
        # 生成輸出文件路徑
        if not output_path:
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            filename = f"survey_{survey_id}_{timestamp}.csv"
            output_path = os.path.join(self.survey_manager.storage_path, "exports", filename)
        
        # 確保目錄存在
        os.makedirs(os.path.dirname(output_path), exist_ok=True)
        
        try:
            with open(output_path, 'w', newline='', encoding='utf-8-sig') as csvfile:
                # 構建標題行
                fieldnames = self._build_csv_headers(survey)
                writer = csv.DictWriter(csvfile, fieldnames=fieldnames)
                writer.writeheader()
                
                # 寫入數據行
                for response in responses:
                    row = self._build_csv_row(survey, response)
                    writer.writerow(row)
            
            print(f"✓ CSV匯出成功: {output_path}")
            return output_path
            
        except Exception as e:
            print(f"CSV匯出失敗: {e}")
            raise
    
    def _build_csv_headers(self, survey: Survey) -> List[str]:
        """構建CSV標題行"""
        headers = [
            '回應ID',
            '受訪者',
            '開始時間', 
            '完成時間',
            '狀態'
        ]
        
        # 添加問題標題
        for question in survey.questions:
            question_header = f"Q{question['id']}: {question['text']}"
            headers.append(question_header)
            
            # 如果是多選題，添加詳細選項欄位
            if question['type'] == 'multiple_choice' and question.get('options'):
                for option in question['options']:
                    headers.append(f"Q{question['id']}_{option}")
        
        return headers
    
    def _build_csv_row(self, survey: Survey, response: SurveyResponse) -> Dict[str, Any]:
        """構建CSV數據行"""
        row = {
            '回應ID': response.response_id,
            '受訪者': response.respondent_name,
            '開始時間': response.started_at,
            '完成時間': response.completed_at or '未完成',
            '狀態': '已完成' if response.is_completed() else '進行中'
        }
        
        # 添加問題回應
        for question in survey.questions:
            question_id = str(question['id'])
            question_header = f"Q{question['id']}: {question['text']}"
            answer = response.responses.get(question_id, '未回答')
            
            if question['type'] == 'multiple_choice':
                # 多選題處理
                if isinstance(answer, list):
                    # 主欄位顯示所有選中項目
                    row[question_header] = ', '.join(str(a) for a in answer)
                    
                    # 為每個選項創建布爾欄位
                    for option in question.get('options', []):
                        option_header = f"Q{question['id']}_{option}"
                        row[option_header] = '1' if option in answer else '0'
                else:
                    row[question_header] = str(answer)
                    # 設置所有選項欄位為0
                    for option in question.get('options', []):
                        option_header = f"Q{question['id']}_{option}"
                        row[option_header] = '0'
            else:
                # 其他題型直接處理
                if isinstance(answer, list):
                    row[question_header] = ', '.join(str(a) for a in answer)
                else:
                    row[question_header] = str(answer)
        
        return row


class JSONExporter(BaseExporter):
    """JSON格式匯出器"""
    
    def export(self, survey_id: str, output_path: str = None) -> str:
        """匯出問卷回應為JSON格式"""
        survey = self.survey_manager.load_survey(survey_id)
        if not survey:
            raise ValueError(f"問卷 {survey_id} 不存在")
        
        responses = self.survey_manager.get_responses_by_survey(survey_id)
        
        # 生成輸出文件路徑
        if not output_path:
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            filename = f"survey_{survey_id}_{timestamp}.json"
            output_path = os.path.join(self.survey_manager.storage_path, "exports", filename)
        
        # 確保目錄存在
        os.makedirs(os.path.dirname(output_path), exist_ok=True)
        
        try:
            export_data = {
                "export_info": {
                    "survey_id": survey_id,
                    "survey_title": survey.title,
                    "export_time": datetime.now().isoformat(),
                    "total_responses": len(responses),
                    "completed_responses": sum(1 for r in responses if r.is_completed())
                },
                "survey": survey.to_dict(),
                "responses": [response.to_dict() for response in responses]
            }
            
            with open(output_path, 'w', encoding='utf-8') as f:
                json.dump(export_data, f, ensure_ascii=False, indent=2)
            
            print(f"✓ JSON匯出成功: {output_path}")
            return output_path
            
        except Exception as e:
            print(f"JSON匯出失敗: {e}")
            raise


class ExcelExporter(BaseExporter):
    """Excel格式匯出器（需要安裝openpyxl）"""
    
    def export(self, survey_id: str, output_path: str = None) -> str:
        """匯出問卷回應為Excel格式"""
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
            from openpyxl.utils.dataframe import dataframe_to_rows
        except ImportError:
            raise ImportError("需要安裝 openpyxl 套件才能匯出Excel格式: pip install openpyxl")
        
        survey = self.survey_manager.load_survey(survey_id)
        if not survey:
            raise ValueError(f"問卷 {survey_id} 不存在")
        
        responses = self.survey_manager.get_responses_by_survey(survey_id)
        if not responses:
            raise ValueError(f"問卷 {survey_id} 沒有回應數據")
        
        # 生成輸出文件路徑
        if not output_path:
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            filename = f"survey_{survey_id}_{timestamp}.xlsx"
            output_path = os.path.join(self.survey_manager.storage_path, "exports", filename)
        
        # 確保目錄存在
        os.makedirs(os.path.dirname(output_path), exist_ok=True)
        
        try:
            # 創建工作簿
            wb = openpyxl.Workbook()
            
            # 創建回應數據工作表
            ws_responses = wb.active
            ws_responses.title = "問卷回應"
            
            # 添加標題和數據
            headers = self._build_excel_headers(survey)
            ws_responses.append(headers)
            
            # 設置標題樣式
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            
            for col_num, header in enumerate(headers, 1):
                cell = ws_responses.cell(row=1, column=col_num)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center", vertical="center")
            
            # 添加數據行
            for response in responses:
                row_data = self._build_excel_row(survey, response)
                ws_responses.append(row_data)
            
            # 自動調整列寬
            for column in ws_responses.columns:
                max_length = 0
                column_letter = column[0].column_letter
                
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                
                adjusted_width = min(max_length + 2, 50)
                ws_responses.column_dimensions[column_letter].width = adjusted_width
            
            # 創建統計摘要工作表
            ws_summary = wb.create_sheet("統計摘要")
            self._add_summary_sheet(ws_summary, survey, responses)
            
            # 保存文件
            wb.save(output_path)
            print(f"✓ Excel匯出成功: {output_path}")
            return output_path
            
        except Exception as e:
            print(f"Excel匯出失敗: {e}")
            raise
    
    def _build_excel_headers(self, survey: Survey) -> List[str]:
        """構建Excel標題行"""
        return [
            '回應ID', '受訪者', '開始時間', '完成時間', '狀態'
        ] + [f"Q{q['id']}: {q['text']}" for q in survey.questions]
    
    def _build_excel_row(self, survey: Survey, response: SurveyResponse) -> List[Any]:
        """構建Excel數據行"""
        row = [
            response.response_id,
            response.respondent_name,
            response.started_at,
            response.completed_at or '未完成',
            '已完成' if response.is_completed() else '進行中'
        ]
        
        # 添加問題回應
        for question in survey.questions:
            question_id = str(question['id'])
            answer = response.responses.get(question_id, '未回答')
            
            if isinstance(answer, list):
                row.append(', '.join(str(a) for a in answer))
            else:
                row.append(str(answer))
        
        return row
    
    def _add_summary_sheet(self, worksheet, survey: Survey, responses: List[SurveyResponse]):
        """添加統計摘要工作表"""
        worksheet.append(["問卷統計摘要"])
        worksheet.append([])
        worksheet.append(["問卷標題", survey.title])
        worksheet.append(["問卷描述", survey.description])
        worksheet.append(["總問題數", len(survey.questions)])
        worksheet.append(["總回應數", len(responses)])
        worksheet.append(["完成回應數", sum(1 for r in responses if r.is_completed())])
        worksheet.append(["完成率", f"{sum(1 for r in responses if r.is_completed()) / len(responses) * 100:.1f}%" if responses else "0%"])
        worksheet.append([])
        
        # 問題統計
        worksheet.append(["問題統計"])
        worksheet.append(["問題編號", "問題類型", "問題內容", "回應率"])
        
        for question in survey.questions:
            question_id = str(question['id'])
            answered_count = sum(1 for r in responses if question_id in r.responses and r.responses[question_id] != '未回答')
            response_rate = f"{answered_count / len(responses) * 100:.1f}%" if responses else "0%"
            
            worksheet.append([
                f"Q{question['id']}",
                question['type'],
                question['text'],
                response_rate
            ])


class SurveyExportManager:
    """問卷匯出管理器"""
    
    def __init__(self, survey_manager: SurveyManager):
        self.survey_manager = survey_manager
        self.exporters = {
            'csv': CSVExporter(survey_manager),
            'json': JSONExporter(survey_manager),
            'excel': ExcelExporter(survey_manager)
        }
    
    def export_survey(self, survey_id: str, format_type: str = 'csv', 
                     output_path: str = None) -> str:
        """統一的問卷匯出接口"""
        if format_type not in self.exporters:
            raise ValueError(f"不支援的匯出格式: {format_type}")
        
        exporter = self.exporters[format_type]
        return exporter.export(survey_id, output_path)
    
    def export_all_formats(self, survey_id: str) -> Dict[str, str]:
        """匯出所有支援的格式"""
        results = {}
        
        for format_type in ['csv', 'json']:  # Excel可能需要額外安裝
            try:
                output_path = self.export_survey(survey_id, format_type)
                results[format_type] = output_path
            except Exception as e:
                print(f"{format_type}匯出失敗: {e}")
                results[format_type] = None
        
        # 嘗試Excel匯出
        try:
            output_path = self.export_survey(survey_id, 'excel')
            results['excel'] = output_path
        except ImportError:
            print("Excel匯出需要安裝openpyxl套件")
            results['excel'] = None
        except Exception as e:
            print(f"Excel匯出失敗: {e}")
            results['excel'] = None
        
        return results
    
    def get_export_files(self) -> List[Dict[str, Any]]:
        """獲取所有匯出文件列表"""
        exports_dir = os.path.join(self.survey_manager.storage_path, "exports")
        if not os.path.exists(exports_dir):
            return []
        
        files = []
        for filename in os.listdir(exports_dir):
            if filename.endswith(('.csv', '.json', '.xlsx')):
                file_path = os.path.join(exports_dir, filename)
                stat = os.stat(file_path)
                
                files.append({
                    'filename': filename,
                    'path': file_path,
                    'size': stat.st_size,
                    'created_time': datetime.fromtimestamp(stat.st_ctime).isoformat(),
                    'format': filename.split('.')[-1]
                })
        
        # 按創建時間排序（新的在前）
        files.sort(key=lambda x: x['created_time'], reverse=True)
        return files